"""General-purpose .xlsx spreadsheet generator using openpyxl.

Usage:
    python create_spreadsheet.py [--output PATH] [--title TEXT] [--subtitle TEXT] [--color HEXCODE]

Example:
    python create_spreadsheet.py --title "Project Report" --color "#2E86AB" --output report.xlsx

Styling defaults (mirrors scripts/create_document.py):
    - Font: 맑은 고딕 (Malgun Gothic)
    - Title: 24pt bold, main color
    - Table header row: 12pt bold white text on main color fill
    - Body cells: 10pt (quote-type columns rendered italic)
    - Main color customizable via --color

Extra features:
    - Column A gets an auto-incrementing "ID" column (skipped if an
      identifier column such as ID/No/번호 already exists).
    - Columns whose header marks them as quotes (quote/quotation/인용구/인용문)
      are rendered in italic.
    - Numeric "growth rate" / "comparison" columns (e.g. 연도별 성장률,
      브랜드별 수치 비교, 증감률, ...) are auto-detected and charted with
      openpyxl.chart (LineChart for time-series categories, BarChart
      otherwise). No chart is created when there isn't enough real
      numeric data.
"""

import argparse
import re
from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONT_NAME = "맑은 고딕"
TITLE_SIZE = 24
SUBTITLE_SIZE = 11
HEADER_SIZE = 12
BODY_SIZE = 10
DEFAULT_COLOR = "1F4E79"

HEX_LENGTH = 6

ID_HEADER = "ID"
IDENTIFIER_KEYWORDS = {"id", "no", "no.", "번호"}
QUOTE_KEYWORDS = {"quote", "quotation", "인용구", "인용문"}
GROWTH_KEYWORDS = ("성장률", "증감률", "증가율", "감소율", "비율 변화", "수치 비교", "변화율")

YEAR_PATTERN = re.compile(r"^\d{4}(년)?$")
MONTH_PATTERN = re.compile(r"^\d{1,2}월$")
QUARTER_PATTERN = re.compile(r"^(제)?\d분기$|^Q\d$", re.IGNORECASE)

MIN_CHART_POINTS = 2
MAX_CHARTS_ON_MAIN_SHEET = 3
CHART_ROW_SPACING = 18


def parse_color(value: str) -> str:
    hex_value = value.lstrip("#")
    if len(hex_value) != HEX_LENGTH or any(
        c not in "0123456789ABCDEFabcdef" for c in hex_value
    ):
        raise argparse.ArgumentTypeError(
            f"'{value}' is not a valid hex color (expected format: RRGGBB or #RRGGBB)"
        )
    return hex_value.upper()


def parse_percentage(value):
    """Safely parse a number or a percentage string (e.g. '12.5%') to float.

    Returns None if the value cannot be interpreted as a number (strings,
    URLs, descriptions, blanks, etc. all fall through to None).
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return None
        if text.endswith("%"):
            text = text[:-1].strip()
        try:
            return float(text)
        except ValueError:
            return None
    return None


def has_identifier_column(headers) -> bool:
    if not headers:
        return False
    return str(headers[0]).strip().lower() in IDENTIFIER_KEYWORDS


def add_id_column(headers, rows):
    """Prepend a sequential 'ID' column, unless an identifier column
    (ID/No/No./번호) is already present at the front of the data.
    """
    if has_identifier_column(headers):
        return list(headers), [list(row) for row in rows]
    new_headers = [ID_HEADER] + list(headers)
    new_rows = [[row_number] + list(row) for row_number, row in enumerate(rows, start=1)]
    return new_headers, new_rows


def is_quote_header(header) -> bool:
    return str(header).strip().lower() in QUOTE_KEYWORDS


def apply_quote_style(cell):
    font = copy(cell.font)
    font.italic = True
    cell.font = font


def _is_timeseries(categories) -> bool:
    if not categories:
        return False
    values = [str(c).strip() for c in categories]
    matches = sum(
        1
        for v in values
        if YEAR_PATTERN.match(v) or MONTH_PATTERN.match(v) or QUARTER_PATTERN.match(v)
    )
    return matches >= max(1, len(values) // 2)


def detect_numeric_comparison(headers, rows):
    """Find columns with enough real numeric data to be worth charting.

    Returns a list of dicts describing each chartable column:
    {col_index (1-based), header, category_col (1-based), is_percentage}.
    Text, URLs, descriptions, blanks, and columns with fewer than
    MIN_CHART_POINTS numeric values are skipped.
    """
    if not headers or not rows:
        return []

    category_col = 2 if has_identifier_column(headers) else 1
    results = []
    for col_index in range(category_col + 1, len(headers) + 1):
        header_name = str(headers[col_index - 1])
        numeric_values = []
        percent_hits = 0
        for row in rows:
            if col_index - 1 >= len(row):
                continue
            raw_value = row[col_index - 1]
            parsed = parse_percentage(raw_value)
            if parsed is not None:
                numeric_values.append(parsed)
                if isinstance(raw_value, str) and raw_value.strip().endswith("%"):
                    percent_hits += 1
        if len(numeric_values) >= MIN_CHART_POINTS:
            is_percentage = percent_hits > 0 or any(
                keyword in header_name for keyword in GROWTH_KEYWORDS
            )
            results.append(
                {
                    "col_index": col_index,
                    "header": header_name,
                    "category_col": category_col,
                    "is_percentage": is_percentage,
                }
            )
    return results


def create_chart(source_ws, chart_ws, header_row, data_start_row, data_end_row, comparison, categories, anchor_cell):
    """Build a Line/Bar chart for one detected numeric comparison column."""
    is_timeseries = _is_timeseries(categories)
    chart = LineChart() if is_timeseries else BarChart()
    chart.title = f"{comparison['header']} 비교"
    chart.y_axis.title = f"{comparison['header']} (%)" if comparison["is_percentage"] else comparison["header"]
    chart.x_axis.title = (
        source_ws.cell(row=header_row, column=comparison["category_col"]).value or "구분"
    )
    chart.width = 15
    chart.height = 8
    chart.style = 10

    data_ref = Reference(
        source_ws,
        min_col=comparison["col_index"],
        max_col=comparison["col_index"],
        min_row=header_row,
        max_row=data_end_row,
    )
    cats_ref = Reference(
        source_ws,
        min_col=comparison["category_col"],
        max_col=comparison["category_col"],
        min_row=data_start_row,
        max_row=data_end_row,
    )
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.legend.position = "b"
    chart_ws.add_chart(chart, anchor_cell)
    return chart


def add_comparison_charts(wb, ws, headers, data_rows, header_row, last_col):
    """Detect growth-rate / numeric-comparison columns and chart them."""
    comparisons = detect_numeric_comparison(headers, data_rows)
    if not comparisons:
        return

    data_start_row = header_row + 1
    data_end_row = header_row + len(data_rows)
    use_separate_sheet = len(comparisons) > MAX_CHARTS_ON_MAIN_SHEET
    chart_ws = wb.create_sheet("Charts") if use_separate_sheet else ws
    anchor_col = 1 if use_separate_sheet else last_col + 2

    for chart_index, comparison in enumerate(comparisons):
        categories = [row[comparison["category_col"] - 1] for row in data_rows]
        anchor_row = header_row + chart_index * CHART_ROW_SPACING
        anchor_cell = f"{get_column_letter(anchor_col)}{anchor_row}"
        create_chart(ws, chart_ws, header_row, data_start_row, data_end_row, comparison, categories, anchor_cell)


def build_workbook(title_text: str, subtitle_text: str, main_color: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    raw_headers = ["Item", "Description", "Value"]
    raw_rows = [
        ("Item 1", "Description of item 1", "-"),
        ("Item 2", "Description of item 2", "-"),
        ("Item 3", "Description of item 3", "-"),
    ]
    headers, sample_rows = add_id_column(raw_headers, raw_rows)
    last_col = len(headers)
    last_col_letter = get_column_letter(last_col)

    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = Font(name=FONT_NAME, size=TITLE_SIZE, bold=True, color=main_color)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells(f"A2:{last_col_letter}2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = subtitle_text
    subtitle_cell.font = Font(name=FONT_NAME, size=SUBTITLE_SIZE, italic=True, color="595959")
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

    header_row = 4
    header_fill = PatternFill(start_color=main_color, end_color=main_color, fill_type="solid")
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_index, value=header)
        cell.font = Font(name=FONT_NAME, size=HEADER_SIZE, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_offset, row_data in enumerate(sample_rows, start=1):
        row_index = header_row + row_offset
        for col_index, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.font = Font(name=FONT_NAME, size=BODY_SIZE)
            header_name = headers[col_index - 1] if col_index - 1 < len(headers) else ""
            if is_quote_header(header_name):
                apply_quote_style(cell)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    for col_index in range(1, last_col + 1):
        column_letter = get_column_letter(col_index)
        max_length = max(
            len(str(ws.cell(row=r, column=col_index).value or ""))
            for r in range(header_row, header_row + len(sample_rows) + 1)
        )
        ws.column_dimensions[column_letter].width = max(12, max_length + 4)

    add_comparison_charts(wb, ws, headers, sample_rows, header_row, last_col)

    return wb


def parse_args():
    parser = argparse.ArgumentParser(description="Generate a general-purpose .xlsx spreadsheet.")
    parser.add_argument("--output", type=Path, default=Path("output.xlsx"), help="Output file path")
    parser.add_argument("--title", default="Spreadsheet Title", help="Spreadsheet title text")
    parser.add_argument(
        "--subtitle",
        default="Subtitle or short description goes here.",
        help="Spreadsheet subtitle text",
    )
    parser.add_argument(
        "--color",
        type=parse_color,
        default=parse_color(DEFAULT_COLOR),
        help="Main color as hex code, e.g. #2E86AB (default: #1F4E79)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook(args.title, args.subtitle, args.color)
    workbook.save(args.output)
    print(f"Spreadsheet saved to {args.output.resolve()}")


if __name__ == "__main__":
    main()
